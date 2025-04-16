import pandas as pd
from pathlib import Path
from typing import List, Union
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def read_file(file_path: str) -> pd.DataFrame:
    """Read a CSV or Excel file into a pandas DataFrame"""
    file_path = Path(file_path)
    if file_path.suffix.lower() == ".csv":
        return pd.read_csv(file_path)
    elif file_path.suffix.lower() in [".xlsx", ".xls"]:
        return pd.read_excel(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path.suffix}")


def merge_files(file_paths: List[str], output_path: str) -> None:
    """Merge multiple CSV or Excel files into one"""
    if not file_paths:
        raise ValueError("No files provided to merge")

    # Determine file type from first file
    first_file = Path(file_paths[0])
    file_type = first_file.suffix.lower()

    # Read all files
    dfs = [read_file(f) for f in file_paths]
    merged_df = pd.concat(dfs, ignore_index=True)

    # Save merged file
    if file_type == ".csv":
        merged_df.to_csv(output_path, index=False)
    else:
        merged_df.to_excel(output_path, index=False)
        format_excel_file(output_path)


def format_excel_file(file_path: str) -> None:
    """Format the Excel file for better readability"""
    wb = load_workbook(file_path)
    ws = wb.active

    # Set header style
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column width
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].auto_size = True

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(file_path)
